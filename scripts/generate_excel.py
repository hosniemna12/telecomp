import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

json_path  = sys.argv[1]
excel_path = sys.argv[2]

with open(json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

GOLD_HEX  = 'C9A84C'
DARK_HEX  = '0A0D14'
GRAY_HEX  = 'F4F6FB'
GREEN_HEX = '22C55E'
RED_HEX   = 'EF4444'
LGRAY_HEX = '8892A4'
WHITE_HEX = 'FFFFFF'

def hdr(ws, row, col, val, bold=True, bg=DARK_HEX, fg=GOLD_HEX, size=10, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, color=fg, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return c

def cell(ws, row, col, val, bold=False, bg=WHITE_HEX, fg='000000', align='left', fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, color=fg, size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if fmt:
        c.number_format = fmt
    return c

thin = Side(style='thin', color='E2E8F0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in row:
            c.border = border

wb = Workbook()

# ═══════════════════════════════════════════════════
# FEUILLE 1 — TABLEAU DE BORD
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Tableau de bord"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 45
ws1.row_dimensions[2].height = 25

# Titre principal
ws1.merge_cells('A1:H1')
c = ws1['A1']
c.value = f"BTL — Rapport de Télécompensation SIBTEL ({data['periode']['type'].title()})"
c.font = Font(name='Arial', bold=True, size=16, color=WHITE_HEX)
c.fill = PatternFill('solid', fgColor=DARK_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A2:H2')
c2 = ws1['A2']
c2.value = f"Période : {data['periode']['debut']} — {data['periode']['fin']}  |  Généré le {data['genere_le']} par {data['genere_par']}"
c2.font = Font(name='Arial', size=9, color=LGRAY_HEX, italic=True)
c2.fill = PatternFill('solid', fgColor=DARK_HEX)
c2.alignment = Alignment(horizontal='center', vertical='center')

# Ligne dorée décorative
ws1.row_dimensions[3].height = 4
for col in range(1, 9):
    ws1.cell(row=3, column=col).fill = PatternFill('solid', fgColor=GOLD_HEX)

# KPIs
ws1.row_dimensions[4].height = 20
ws1.row_dimensions[5].height = 18
stats = data['stats']

hdr(ws1, 4, 1, 'INDICATEUR', bg=DARK_HEX, fg=GOLD_HEX, size=9)
hdr(ws1, 4, 2, 'VALEUR', bg=DARK_HEX, fg=GOLD_HEX, size=9)

kpis = [
    ('Total fichiers reçus',            stats['total_fichiers'],     False),
    ('Fichiers traités avec succès',    stats['total_traites'],      False),
    ('Fichiers en erreur',              stats['total_erreurs'],       True),
    ('Fichiers en attente validation',  stats['total_en_attente'],    False),
    ('Total transactions',              stats['total_transactions'],  False),
    ('Transactions valides',            stats['total_valides'],       False),
    ('Transactions rejetées',           stats['total_rejetes'],       True),
    ('Montant total validé (TND)',       stats['montant_total'],       False),
    ('Total rejets SIBTEL',             stats['total_rejets'],        True),
]

for i, (label, val, is_alert) in enumerate(kpis):
    row = 5 + i
    ws1.row_dimensions[row].height = 18
    bg = 'FFF5F5' if is_alert and val > 0 else (GRAY_HEX if i % 2 == 0 else WHITE_HEX)
    fg = RED_HEX if is_alert and val > 0 else '000000'
    cell(ws1, row, 1, label, bg=bg)
    if label == 'Montant total validé (TND)':
        c = cell(ws1, row, 2, val, bold=True, bg=bg, fg='1a5e38', align='right', fmt='#,##0.000')
    else:
        cell(ws1, row, 2, val, bold=True, bg=bg, fg=fg, align='right')

apply_border(ws1, 4, 4+len(kpis), 1, 2)

# Répartition par type
row_start = 5 + len(kpis) + 2
ws1.row_dimensions[row_start].height = 5
ws1.row_dimensions[row_start+1].height = 20
hdr(ws1, row_start+1, 4, 'TYPE', bg=DARK_HEX, fg=GOLD_HEX, size=9)
hdr(ws1, row_start+1, 5, 'FICHIERS', bg=DARK_HEX, fg=GOLD_HEX, size=9)
hdr(ws1, row_start+1, 6, 'MONTANT (TND)', bg=DARK_HEX, fg=GOLD_HEX, size=9)

for j, (k, v) in enumerate(data.get('par_type', {}).items()):
    r = row_start + 2 + j
    ws1.row_dimensions[r].height = 16
    bg = GRAY_HEX if j % 2 == 0 else WHITE_HEX
    cell(ws1, r, 4, v['type'], bg=bg)
    cell(ws1, r, 5, v['count'], bg=bg, align='right')
    cell(ws1, r, 6, v['montant'], bg=bg, align='right', fmt='#,##0.000')

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 5
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 18

# ═══════════════════════════════════════════════════
# FEUILLE 2 — LISTING FICHIERS
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Fichiers")
ws2.sheet_view.showGridLines = False
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A1:I1')
c = ws2['A1']
c.value = "Listing des fichiers traités"
c.font = Font(name='Arial', bold=True, size=13, color=WHITE_HEX)
c.fill = PatternFill('solid', fgColor=DARK_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')

headers = ['Nom fichier', 'Type', 'Statut', 'Transactions', 'Rejets', 'Montant (TND)', 'Date', 'Uploadé par', 'Validé par']
for col, h in enumerate(headers, 1):
    ws2.row_dimensions[2].height = 20
    hdr(ws2, 2, col, h, size=9)

for i, f in enumerate(data['fichiers']):
    row = 3 + i
    ws2.row_dimensions[row].height = 16
    bg = GRAY_HEX if i % 2 == 0 else WHITE_HEX
    is_ok = f['statut'] in ['TRAITE', 'VALIDE']
    is_err = f['statut'] == 'ERREUR'
    if is_err:
        bg = 'FFF5F5'
    cell(ws2, row, 1, f['nom'], bg=bg)
    cell(ws2, row, 2, f['type'], bg=bg, align='center')
    fg_s = GREEN_HEX if is_ok else (RED_HEX if is_err else 'f59e0b')
    cell(ws2, row, 3, f['statut'], bg=bg, fg=fg_s, bold=True, align='center')
    cell(ws2, row, 4, f['transactions'], bg=bg, align='right')
    cell(ws2, row, 5, f['rejets'], bg=bg, align='right', fg=RED_HEX if f['rejets'] > 0 else '000000')
    cell(ws2, row, 6, f['montant'] if isinstance(f['montant'], (int,float)) else 0, bg=bg, align='right', fmt='#,##0.000')
    cell(ws2, row, 7, f['date'], bg=bg, align='center')
    cell(ws2, row, 8, f['uploade_par'], bg=bg)
    cell(ws2, row, 9, f['valide_par'], bg=bg)

apply_border(ws2, 2, 2+len(data['fichiers']), 1, 9)

cols_w = [40, 12, 22, 13, 8, 16, 16, 18, 18]
for i, w in enumerate(cols_w, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════
# FEUILLE 3 — TRANSACTIONS
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Transactions")
ws3.sheet_view.showGridLines = False
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A1:H1')
c = ws3['A1']
c.value = "Listing des transactions"
c.font = Font(name='Arial', bold=True, size=13, color=WHITE_HEX)
c.fill = PatternFill('solid', fgColor=DARK_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')

hdrs3 = ['ID Fichier', 'RIB Donneur', 'Nom Donneur', 'RIB Bénéficiaire', 'Nom Bénéficiaire', 'Montant (TND)', 'Statut', 'Motif']
for col, h in enumerate(hdrs3, 1):
    ws3.row_dimensions[2].height = 20
    hdr(ws3, 2, col, h, size=9)

for i, d in enumerate(data.get('details', [])):
    row = 3 + i
    ws3.row_dimensions[row].height = 15
    bg = GRAY_HEX if i % 2 == 0 else WHITE_HEX
    is_ok = d['statut'] == 'VALIDE'
    cell(ws3, row, 1, d['fichier_id'], bg=bg, align='center')
    cell(ws3, row, 2, d['rib_donneur'], bg=bg, fmt='@')
    cell(ws3, row, 3, d['nom_donneur'], bg=bg)
    cell(ws3, row, 4, d['rib_beneficiaire'], bg=bg, fmt='@')
    cell(ws3, row, 5, d['nom_beneficiaire'], bg=bg)
    cell(ws3, row, 6, d['montant'], bg=bg, align='right', fmt='#,##0.000')
    cell(ws3, row, 7, d['statut'], bg=bg, fg=GREEN_HEX if is_ok else RED_HEX, bold=True, align='center')
    cell(ws3, row, 8, d['motif_operation'], bg=bg)

if data.get('details'):
    apply_border(ws3, 2, 2+len(data['details']), 1, 8)

cols_w3 = [10, 22, 25, 22, 25, 16, 12, 30]
for i, w in enumerate(cols_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════
# FEUILLE 4 — REJETS
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Rejets")
ws4.sheet_view.showGridLines = False
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A1:D1')
c = ws4['A1']
c.value = "Récapitulatif des rejets"
c.font = Font(name='Arial', bold=True, size=13, color=WHITE_HEX)
c.fill = PatternFill('solid', fgColor='B91C1C')
c.alignment = Alignment(horizontal='center', vertical='center')

hdrs4 = ['Fichier', 'Code rejet', 'Motif', 'Étape détection']
for col, h in enumerate(hdrs4, 1):
    hdr(ws4, 2, col, h, bg='B91C1C', fg=WHITE_HEX, size=9)

for i, r in enumerate(data.get('rejets', [])):
    row = 3 + i
    ws4.row_dimensions[row].height = 15
    bg = 'FFF5F5' if i % 2 == 0 else WHITE_HEX
    cell(ws4, row, 1, r['fichier'], bg=bg)
    cell(ws4, row, 2, r['code'], bg=bg, fg=RED_HEX, bold=True, align='center')
    cell(ws4, row, 3, r['motif'], bg=bg)
    cell(ws4, row, 4, r['etape'], bg=bg, align='center')

if data.get('rejets'):
    apply_border(ws4, 2, 2+len(data['rejets']), 1, 4)

cols_w4 = [40, 15, 50, 20]
for i, w in enumerate(cols_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

wb.save(excel_path)
print("Excel généré avec succès")
